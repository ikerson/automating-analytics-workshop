import matplotlib.pyplot as plt
import pandas as pd
from pathlib import Path


def save_top_schools_chart(top_schools_df, output_dir):
    path = Path(output_dir) / "top_middle_schools.png"
    fig, ax = plt.subplots(figsize=(10, 6))
    ax.barh(top_schools_df['middle_school_name'], top_schools_df['student_count'], color='steelblue')
    ax.set_xlabel("Number of Students")
    ax.set_title("Top 10 Middle Schools by Student Count")
    ax.invert_yaxis()
    plt.tight_layout()
    plt.savefig(path)
    plt.close()
    return path


def save_size_chart(size_df, output_dir):
    path = Path(output_dir) / "school_size_distribution.png"
    fig, ax = plt.subplots(figsize=(7, 5))
    ax.bar(size_df['school_size'].astype(str), size_df['student_count'], color='teal')
    ax.set_xlabel("School Size")
    ax.set_ylabel("Number of Students")
    ax.set_title("Students by Middle School Size")
    plt.tight_layout()
    plt.savefig(path)
    plt.close()
    return path


def save_excel_report(merged_df, top_schools_df, zip_summary_df, size_df, chart_paths, output_dir):
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as XLImage

    path = Path(output_dir) / "student_report.xlsx"
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        merged_df.to_excel(writer, sheet_name='Student Data', index=False)
        top_schools_df.to_excel(writer, sheet_name='Top 10 Schools', index=False)
        zip_summary_df.to_excel(writer, sheet_name='By ZIP', index=False)
        size_df.to_excel(writer, sheet_name='By School Size', index=False)

    wb = load_workbook(path)
    ws = wb.create_sheet('Charts')
    row = 1
    for chart_path in chart_paths:
        img = XLImage(str(chart_path))
        ws.add_image(img, f'A{row}')
        row += 30
    wb.save(path)
    return path
