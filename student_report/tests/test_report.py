import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

import matplotlib
matplotlib.use('Agg')

import pandas as pd
import pytest
from report import save_top_schools_chart, save_size_chart, save_excel_report


@pytest.fixture
def top_schools_df():
    return pd.DataFrame({
        'middle_school_name': ['IS 71', 'Hoboken Middle School'],
        'city_location': ['New York', 'Hoboken'],
        'zip_mailing': ['10001', '07030'],
        'school_size': ['Medium (300-700)', 'Small (<300)'],
        'student_count': [5, 3],
        'school_enrollment': [450, 250],
    })


@pytest.fixture
def size_df():
    return pd.DataFrame({
        'school_size': ['Small (<300)', 'Medium (300-700)', 'Large (700+)'],
        'student_count': [3, 5, 0],
    })


@pytest.fixture
def merged_df():
    return pd.DataFrame({
        'student_id': [1, 2],
        'first_name': ['Alice', 'Bob'],
        'middle_school_name': ['IS 71', 'Hoboken Middle School'],
        'ncessch': ['360007702472', '340183001982'],
    })


@pytest.fixture
def zip_summary_df():
    return pd.DataFrame({
        'zip_mailing': ['10001', '07030'],
        'city_location': ['New York', 'Hoboken'],
        'student_count': [5, 3],
    })


def test_save_top_schools_chart(top_schools_df, tmp_path):
    path = save_top_schools_chart(top_schools_df, tmp_path)
    assert path.exists()
    assert path.suffix == '.png'


def test_save_size_chart(size_df, tmp_path):
    path = save_size_chart(size_df, tmp_path)
    assert path.exists()
    assert path.suffix == '.png'


def test_save_excel_report_sheets(merged_df, top_schools_df, zip_summary_df, size_df, tmp_path):
    from openpyxl import load_workbook
    chart1 = save_top_schools_chart(top_schools_df, tmp_path)
    chart2 = save_size_chart(size_df, tmp_path)
    path = save_excel_report(merged_df, top_schools_df, zip_summary_df, size_df, [chart1, chart2], tmp_path)
    wb = load_workbook(path)
    assert wb.sheetnames == ['Student Data', 'Top 10 Schools', 'By ZIP', 'By School Size', 'Charts']
